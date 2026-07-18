"""Admin content portal API. Every route in this router sits behind require_admin (router-level
dependency), so an unauthenticated caller gets 401 and an authenticated non-admin gets 403. This is
the "no outsiders" perimeter for authoring the syllabus graph and the item bank.

What it exposes:
  * syllabus read/write  — exams, sections, topics, concepts, prerequisite edges
  * item lifecycle       — list (with answers), create, edit, approve, retire, delete, bulk xlsx
  * admin management     — list / grant / revoke admins

Bulk question volume still goes through the same importer the /ingest path uses; the portal is for
the things a spreadsheet is bad at (defining the graph, approving/QC, spot-editing single items).
"""
from __future__ import annotations

import io
import uuid
from datetime import datetime
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Response, UploadFile
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from .. import models
from ..config import settings
from ..deps import get_db
from ..schemas import IngestReport, ItemIn
from ..services import html_sanitize, ingestion, knowledge_graph as kg, question_bank, security
from ..services.admin_auth import grant_admin, require_admin, revoke_admin

router = APIRouter(prefix="/admin", tags=["admin"], dependencies=[Depends(require_admin)])


@router.post("/seed-demo-content")
def seed_demo_content(db: Session = Depends(get_db)) -> dict:
    """Dev-only: (re)seed the CAT/GMAT/GRE learning tree (sections -> chapters -> subtopics -> quiz).
    Runs inside the app so it uses the warm connection pool (avoids Neon control-plane throttling on
    fresh connections). Admin-gated + dev_mode-gated."""
    if not settings.dev_mode:
        raise HTTPException(403, "seeding is disabled outside dev_mode")
    from scripts.seed_demo_content import run as _seed_run

    counts = _seed_run(db)
    return {"ok": True, **counts}


# ───────────────────────── whoami ─────────────────────────
@router.get("/me")
def admin_me(admin: models.Account = Depends(require_admin)) -> dict:
    return {"account_id": str(admin.id), "email": admin.email,
            "display_name": admin.display_name, "is_admin": True}


# ───────────────────────── syllabus: read ─────────────────────────
@router.get("/exams")
def list_exams(db: Session = Depends(get_db)) -> list[dict]:
    return [{"code": e.code, "name": e.name} for e in db.scalars(select(models.Exam)).all()]


@router.get("/syllabus")
def syllabus(exam: str, db: Session = Depends(get_db)) -> dict:
    ex = db.get(models.Exam, exam)
    if ex is None:
        raise HTTPException(404, f"no exam {exam!r}")
    sections = db.scalars(select(models.Section).where(models.Section.exam_code == exam)).all()
    nodes = db.scalars(select(models.KnowledgeNode).where(models.KnowledgeNode.exam_code == exam)).all()
    prereqs = db.scalars(select(models.PrereqEdge)).all()
    items = db.scalars(select(models.Item).where(models.Item.exam_code == exam)).all()
    sec_by_id = {s.id: s for s in sections}
    item_count: dict[str, int] = {}
    for it in items:
        item_count[it.concept_node_id] = item_count.get(it.concept_node_id, 0) + 1
    pre_by_node: dict[str, list[str]] = {}
    for e in prereqs:
        pre_by_node.setdefault(e.node_id, []).append(e.prereq_node_id)
    return {
        "exam": {"code": ex.code, "name": ex.name},
        "sections": [{"id": str(s.id), "key": s.key, "name": s.name} for s in sections],
        "nodes": [{
            "id": n.id, "name": n.name, "kind": n.kind, "parent_id": n.parent_id,
            "section": sec_by_id[n.section_id].key if n.section_id in sec_by_id else None,
            "item_count": item_count.get(n.id, 0),
            "prereqs": pre_by_node.get(n.id, []),
        } for n in nodes],
    }


# ───────────────────────── syllabus: write ─────────────────────────
class ExamIn(BaseModel):
    code: str
    name: str


@router.post("/exams")
def create_exam(body: ExamIn, db: Session = Depends(get_db)) -> dict:
    if db.get(models.Exam, body.code):
        raise HTTPException(409, f"exam {body.code!r} already exists")
    db.add(models.Exam(code=body.code, name=body.name))
    db.commit()
    return {"code": body.code, "name": body.name}


class SectionIn(BaseModel):
    exam_code: str
    key: str
    name: str


@router.post("/sections")
def create_section(body: SectionIn, db: Session = Depends(get_db)) -> dict:
    if db.get(models.Exam, body.exam_code) is None:
        raise HTTPException(404, f"no exam {body.exam_code!r}")
    if db.scalar(select(models.Section).where(
            models.Section.exam_code == body.exam_code, models.Section.key == body.key)):
        raise HTTPException(409, f"section {body.key!r} already exists in {body.exam_code}")
    s = models.Section(exam_code=body.exam_code, key=body.key, name=body.name)
    db.add(s)
    db.commit()
    return {"id": str(s.id), "key": s.key, "name": s.name}


class NodeIn(BaseModel):
    id: str
    exam_code: str
    section_key: str
    name: str
    parent_id: Optional[str] = None


def _section_id(db: Session, exam_code: str, section_key: str):
    s = db.scalar(select(models.Section).where(
        models.Section.exam_code == exam_code, models.Section.key == section_key))
    if s is None:
        raise HTTPException(404, f"no section {section_key!r} in {exam_code}")
    return s.id


@router.post("/topics")
def create_topic(body: NodeIn, db: Session = Depends(get_db)) -> dict:
    if db.get(models.KnowledgeNode, body.id):
        raise HTTPException(409, f"node {body.id!r} already exists")
    sid = _section_id(db, body.exam_code, body.section_key)
    n = models.KnowledgeNode(id=body.id, exam_code=body.exam_code, section_id=sid,
                             kind="topic", name=body.name, parent_id=None)
    db.add(n)
    db.commit()
    return {"id": n.id, "name": n.name, "kind": "topic"}


@router.post("/concepts")
def create_concept(body: NodeIn, db: Session = Depends(get_db)) -> dict:
    if db.get(models.KnowledgeNode, body.id):
        raise HTTPException(409, f"node {body.id!r} already exists")
    sid = _section_id(db, body.exam_code, body.section_key)
    if body.parent_id:
        parent = db.get(models.KnowledgeNode, body.parent_id)
        if parent is None or parent.kind != "topic":
            raise HTTPException(400, "parent_id must be an existing topic node")
    n = models.KnowledgeNode(id=body.id, exam_code=body.exam_code, section_id=sid,
                             kind="concept", name=body.name, parent_id=body.parent_id)
    db.add(n)
    db.commit()
    return {"id": n.id, "name": n.name, "kind": "concept", "parent_id": n.parent_id}


class PrereqIn(BaseModel):
    node_id: str
    prereq_node_id: str


@router.post("/prereqs")
def add_prereq(body: PrereqIn, db: Session = Depends(get_db)) -> dict:
    if db.get(models.KnowledgeNode, body.node_id) is None or \
            db.get(models.KnowledgeNode, body.prereq_node_id) is None:
        raise HTTPException(404, "both node_id and prereq_node_id must exist")
    if body.node_id == body.prereq_node_id:
        raise HTTPException(400, "a node cannot be its own prerequisite")
    if db.scalar(select(models.PrereqEdge).where(
            models.PrereqEdge.node_id == body.node_id,
            models.PrereqEdge.prereq_node_id == body.prereq_node_id)):
        return {"ok": True, "already": True}
    db.add(models.PrereqEdge(node_id=body.node_id, prereq_node_id=body.prereq_node_id))
    db.commit()
    return {"ok": True}


@router.delete("/prereqs")
def del_prereq(body: PrereqIn, db: Session = Depends(get_db)) -> dict:
    row = db.scalar(select(models.PrereqEdge).where(
        models.PrereqEdge.node_id == body.node_id,
        models.PrereqEdge.prereq_node_id == body.prereq_node_id))
    if row:
        db.delete(row)
        db.commit()
    return {"ok": True}


@router.delete("/nodes/{node_id}")
def delete_node(node_id: str, cascade: bool = False, db: Session = Depends(get_db)) -> dict:
    """Delete a topic/concept node. Without ``cascade`` the node must be empty (no child concepts,
    no items) — the safe default. With ``cascade=1`` it removes the whole subtree in one shot: every
    descendant concept, all their items (plus those items' responses/exposures), attached materials,
    per-learner state and prerequisite edges. This is what the content portal's "delete chapter /
    subtopic" uses so an admin can remove authored content that already has questions."""
    n = db.get(models.KnowledgeNode, node_id)
    if n is None:
        raise HTTPException(404, f"no node {node_id!r}")

    children = db.scalars(
        select(models.KnowledgeNode).where(models.KnowledgeNode.parent_id == node_id)).all()

    if not cascade:
        if children:
            raise HTTPException(409, "node has child concepts — delete or reparent them first")
        if db.scalar(select(models.Item).where(models.Item.concept_node_id == node_id)):
            raise HTTPException(409, "node has items — delete/retire them first")
        db.execute(delete(models.PrereqEdge).where(
            (models.PrereqEdge.node_id == node_id) | (models.PrereqEdge.prereq_node_id == node_id)))
        db.delete(n)
        db.commit()
        return {"ok": True, "deleted": node_id}

    # cascade: this node + all descendant concepts, deleting dependents in FK-safe order
    node_ids = [node_id] + [c.id for c in children]
    item_ids = list(db.scalars(
        select(models.Item.item_id).where(models.Item.concept_node_id.in_(node_ids))).all())
    if item_ids:
        db.execute(delete(models.Response).where(models.Response.item_id.in_(item_ids)))
        db.execute(delete(models.Exposure).where(models.Exposure.item_id.in_(item_ids)))
        db.execute(delete(models.Item).where(models.Item.item_id.in_(item_ids)))
    db.execute(delete(models.Material).where(models.Material.node_id.in_(node_ids)))
    db.execute(delete(models.LearnerNodeState).where(models.LearnerNodeState.node_id.in_(node_ids)))
    db.execute(delete(models.PrereqEdge).where(
        models.PrereqEdge.node_id.in_(node_ids) | models.PrereqEdge.prereq_node_id.in_(node_ids)))
    for c in children:          # children reference this node via parent_id — remove them first
        db.delete(c)
    db.flush()
    db.delete(n)
    db.commit()
    return {"ok": True, "deleted": node_id, "cascade": True,
            "concepts_removed": len(children), "items_removed": len(item_ids)}


class NodeRenameIn(BaseModel):
    name: str


@router.patch("/nodes/{node_id}")
def rename_node(node_id: str, body: NodeRenameIn, db: Session = Depends(get_db)) -> dict:
    """Rename a chapter (topic) or subtopic (concept). Content, children and items are untouched."""
    n = db.get(models.KnowledgeNode, node_id)
    if n is None:
        raise HTTPException(404, f"no node {node_id!r}")
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(400, "name cannot be empty")
    n.name = name
    db.commit()
    return {"id": n.id, "name": n.name, "kind": n.kind}


@router.post("/nodes/{node_id}/practice-bank")
def ensure_practice_bank(node_id: str, db: Session = Depends(get_db)) -> dict:
    """Get-or-create a hidden practice pool for a SUBTOPIC (or a whole chapter). The questions live in
    ``<node_id>__practice`` — a concept parented to the chapter, so it counts toward the chapter's
    mastery/accuracy but never shows up as a subtopic. The student practice section runs the MAB over
    these per-subtopic pools. Author questions into it via the item endpoints using the returned id."""
    node = db.get(models.KnowledgeNode, node_id)
    if node is None:
        raise HTTPException(404, f"no node {node_id!r}")
    pid = kg.practice_bank_node_id(node_id)
    pool = db.get(models.KnowledgeNode, pid)
    if pool is None:
        # keep the pool a child of the chapter: a subtopic's pool sits under its chapter (node.parent),
        # a chapter's pool under the chapter itself.
        is_sub = node.kind == models.NodeKind.concept.value
        parent_id = node.parent_id if is_sub else node.id
        name = f"{node.name} · practice" if is_sub else "Chapter practice bank"
        pool = models.KnowledgeNode(
            id=pid, exam_code=node.exam_code, section_id=node.section_id,
            kind=models.NodeKind.concept.value, name=name, parent_id=parent_id)
        db.add(pool)
        db.commit()
    return {"id": pid, "node_id": node_id, "name": pool.name}


# ───────────────────────── items ─────────────────────────
@router.get("/items")
def list_items(exam: Optional[str] = None, section: Optional[str] = None,
               concept: Optional[str] = None, status: Optional[str] = None,
               limit: int = 200, db: Session = Depends(get_db)) -> dict:
    q = select(models.Item)
    if exam:
        q = q.where(models.Item.exam_code == exam)
    if concept:
        q = q.where(models.Item.concept_node_id == concept)
    if status:
        q = q.where(models.Item.status == status)
    items = db.scalars(q.limit(limit)).all()
    sec_key = {s.id: s.key for s in db.scalars(select(models.Section)).all()}
    out = []
    for it in items:
        if section and sec_key.get(it.section_id) != section:
            continue
        out.append({
            "item_id": it.item_id, "version": it.version, "exam_code": it.exam_code,
            "section": sec_key.get(it.section_id), "concept_node_id": it.concept_node_id,
            "difficulty_d": it.difficulty_d, "format": it.format, "num_options": it.num_options,
            "stem": it.stem, "options": it.options, "correct_answer": it.correct_answer,
            "solution": it.solution, "status": it.status,
        })
    return {"count": len(out), "items": out}


@router.post("/items", response_model=IngestReport)
def create_item(body: ItemIn, db: Session = Depends(get_db)) -> IngestReport:
    """Create (or upsert) a single item. Goes through the same validated ingest path as bulk upload,
    so the authored-vs-derived boundary (no IRT a/b/c authoring) is enforced identically."""
    return ingestion.ingest_items(db, [body])


class ItemPatch(BaseModel):
    model_config = ConfigDict(extra="forbid")
    stem: Optional[str] = None
    options: Optional[list[str]] = None
    correct_answer: Optional[str] = None
    solution: Optional[str] = None
    difficulty_d: Optional[int] = Field(default=None, ge=-2, le=2)
    status: Optional[str] = None
    negative_marking: Optional[bool] = None


@router.patch("/items/{item_id}")
def edit_item(item_id: str, body: ItemPatch, db: Session = Depends(get_db)) -> dict:
    it = db.get(models.Item, item_id)
    if it is None:
        raise HTTPException(404, f"no item {item_id!r}")
    data = body.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(it, k, v)
    if data:
        it.version = (it.version or 1) + 1
    db.commit()
    return {"ok": True, "item_id": item_id, "version": it.version}


@router.post("/items/{item_id}/approve")
def approve_item(item_id: str, db: Session = Depends(get_db)) -> dict:
    it = db.get(models.Item, item_id)
    if it is None:
        raise HTTPException(404, f"no item {item_id!r}")
    it.status = "approved"
    db.commit()
    return {"ok": True, "item_id": item_id, "status": "approved"}


@router.post("/items/{item_id}/retire")
def retire_item(item_id: str, db: Session = Depends(get_db)) -> dict:
    it = db.get(models.Item, item_id)
    if it is None:
        raise HTTPException(404, f"no item {item_id!r}")
    it.status = "retired"
    db.commit()
    return {"ok": True, "item_id": item_id, "status": "retired"}


@router.delete("/items/{item_id}")
def delete_item(item_id: str, db: Session = Depends(get_db)) -> dict:
    it = db.get(models.Item, item_id)
    if it is None:
        raise HTTPException(404, f"no item {item_id!r}")
    # Remove rows that reference this item first, or the FK blocks the delete and the row
    # survives in the DB while the admin UI has already dropped it (FK-safe order).
    db.execute(delete(models.IrtParameter).where(models.IrtParameter.item_id == item_id))
    db.execute(delete(models.Response).where(models.Response.item_id == item_id))
    db.execute(delete(models.Exposure).where(models.Exposure.item_id == item_id))
    db.delete(it)
    db.commit()
    return {"ok": True, "deleted": item_id}


@router.post("/items/upload-xlsx")
async def upload_xlsx(file: UploadFile = File(...), scope: str = "both", db: Session = Depends(get_db)) -> dict:
    """Admin-gated bulk upload — same question-bank workbook format as the authors' path: one sheet
    per exam (tab name = exam) OR a single sheet with an `Exam` column. Builds the graph from
    Topic/Subtopic/Prerequisites and ingests every question atomically; returns the import report."""
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise HTTPException(500, "openpyxl not installed")
    data = await file.read()
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"could not read workbook: {e}")
    rows: list[dict] = []
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        for r in it:
            if r is None or all(c is None for c in r):
                continue
            row: dict[str, Any] = dict(zip(header, r))
            if not row.get("Exam"):
                row["Exam"] = ws.title
            rows.append(row)
    if not rows:
        raise HTTPException(400, "no data rows found in any sheet")
    return question_bank.import_question_bank(db, rows, usage_scope=scope)


# ───────────────────────── admin management ─────────────────────────
@router.get("/admins")
def list_admins(db: Session = Depends(get_db)) -> list[dict]:
    rows = db.scalars(select(models.AdminUser)).all()
    out = []
    for r in rows:
        acc = db.get(models.Account, r.account_id)
        out.append({"account_id": str(r.account_id), "role": r.role,
                    "email": acc.email if acc else None,
                    "display_name": acc.display_name if acc else None})
    return out


class GrantIn(BaseModel):
    email: str
    password: Optional[str] = None
    display_name: Optional[str] = ""


@router.post("/admins")
def grant(body: GrantIn, db: Session = Depends(get_db)) -> dict:
    """Add an admin. With a password, creates the account (or resets its password) and grants admin —
    so an admin can mint new admins from the portal. Without a password, grants admin to an account
    that already exists (the original behaviour)."""
    email = (body.email or "").strip().lower()
    if not email:
        raise HTTPException(400, "email is required")

    if body.password:
        if len(body.password) < 8:
            raise HTTPException(400, "password must be at least 8 characters")
        pw_hash = security.hash_password(body.password)
        acc = db.scalar(select(models.Account).where(models.Account.email == email))
        if acc is None:
            acc = models.Account(id=uuid.uuid4(), email=email, display_name=(body.display_name or "Admin"))
            db.add(acc)
            db.flush()
            db.add(models.Credential(account_id=acc.id, password_hash=pw_hash))
        else:
            cred = db.get(models.Credential, acc.id)
            if cred is None:
                db.add(models.Credential(account_id=acc.id, password_hash=pw_hash))
            else:
                cred.password_hash = pw_hash
        if db.get(models.AdminUser, acc.id) is None:
            db.add(models.AdminUser(account_id=acc.id, role="admin"))
        db.commit()
        return {"ok": True, "account_id": str(acc.id), "email": acc.email}

    acc = grant_admin(db, email)
    return {"ok": True, "account_id": str(acc.id), "email": acc.email}


@router.delete("/admins/{account_id}")
def revoke(account_id: str, admin: models.Account = Depends(require_admin),
           db: Session = Depends(get_db)) -> dict:
    if str(admin.id) == account_id:
        raise HTTPException(400, "you cannot revoke your own admin access")
    ok = revoke_admin(db, account_id)
    if not ok:
        raise HTTPException(404, "that account is not an admin")
    return {"ok": True, "revoked": account_id}


# ═══════════════════════ learning content (concept theory + videos) ═══════════════════════
class VideoIn(BaseModel):
    title: str
    url: str
    seconds: Optional[int] = None


class ContentIn(BaseModel):
    body: Optional[str] = ""          # concept explanation (markdown/plain text)
    videos: list[VideoIn] = []


@router.get("/concepts/{node_id}/content")
def get_concept_content(node_id: str, db: Session = Depends(get_db)) -> dict:
    """Learning content for a subtopic: the concept explanation + its video links."""
    n = db.get(models.KnowledgeNode, node_id)
    if n is None or n.kind != "concept":
        raise HTTPException(404, f"no concept {node_id!r}")
    t = n.theory or {}
    return {"node_id": node_id, "name": n.name,
            "body": t.get("body", ""), "videos": t.get("videos", [])}


@router.put("/concepts/{node_id}/content")
def set_concept_content(node_id: str, body: ContentIn, db: Session = Depends(get_db)) -> dict:
    n = db.get(models.KnowledgeNode, node_id)
    if n is None or n.kind != "concept":
        raise HTTPException(404, f"no concept {node_id!r}")
    # body may contain HTML — sanitize before storing (it renders into students' browsers).
    n.theory = {"body": html_sanitize.sanitize_html(body.body or ""),
                "videos": [v.model_dump() for v in body.videos]}
    db.commit()
    return {"ok": True, "node_id": node_id, "videos": len(body.videos)}


@router.post("/concepts/{node_id}/content/html")
async def upload_concept_html(node_id: str, file: UploadFile = File(...),
                              db: Session = Depends(get_db)) -> dict:
    """Upload an .html file as a concept's explanation. The file is read, sanitized (scripts, event
    handlers, iframes and javascript: URLs removed), and stored; existing video links are preserved.
    Returns the cleaned HTML so the admin console can preview exactly what students will see.

    Note: images must use full URLs (e.g. https://.../fig.png). Relative paths like <img src="fig.png">
    point at files on your computer and won't load for students — host images and link them by URL.
    """
    n = db.get(models.KnowledgeNode, node_id)
    if n is None or n.kind != "concept":
        raise HTTPException(404, f"no concept {node_id!r}")
    raw = await file.read()
    if len(raw) > 2_000_000:
        raise HTTPException(413, "HTML file too large (max 2 MB).")
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("latin-1", errors="replace")
    clean = html_sanitize.sanitize_html(text)
    existing = n.theory or {}
    n.theory = {"body": clean, "videos": existing.get("videos", [])}
    db.commit()
    return {"ok": True, "node_id": node_id, "name": n.name, "chars": len(clean), "body": clean}


# ═══════════════════════ per-subtopic quiz: bulk upload (simple sheet) ═══════════════════════
_QUIZ_HEADERS = {
    "question id": "id", "id": "id",
    "difficulty (-2 to 2)": "diff", "difficulty": "diff", "d": "diff",
    "question text": "stem", "question": "stem", "stem": "stem",
    "option a": "A", "option b": "B", "option c": "C", "option d": "D", "option e": "E",
    "correct answer": "correct", "answer": "correct", "correct": "correct",
    "solution / explanation": "sol", "solution": "sol", "explanation": "sol",
    "expected time (sec)": "time", "time": "time",
}


def _parse_quiz_xlsx(data: bytes, exam: str, section_key: str, concept_node_id: str, scope: str):
    from openpyxl import load_workbook

    from ..schemas import ItemFormatIn, ItemIn, UsageScopeIn
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"could not read workbook: {e}")
    ws = wb.worksheets[0]
    rit = ws.iter_rows(values_only=True)
    header = next(rit, None)
    if not header:
        raise HTTPException(400, "empty sheet")
    cols: dict[str, int] = {}
    for i, h in enumerate(header):
        key = _QUIZ_HEADERS.get(str(h).strip().lower()) if h is not None else None
        if key:
            cols[key] = i
    if "stem" not in cols or "correct" not in cols:
        raise HTTPException(400, "sheet needs at least a 'Question text' (or 'Question') and a "
                                 "'Correct answer' (or 'Answer') column")
    items, errors = [], []
    seq = 0
    for r in rit:
        if r is None or all(c is None for c in r):
            continue

        def cell(k):
            i = cols.get(k)
            return r[i] if (i is not None and i < len(r)) else None

        stem = cell("stem")
        if stem is None or str(stem).strip() == "":
            continue
        seq += 1
        opts = [str(cell(L)).strip() for L in "ABCDE" if cell(L) not in (None, "")]
        diff = cell("diff")
        try:
            diff = max(-2, min(2, int(float(diff)))) if diff is not None else 0
        except (TypeError, ValueError):
            diff = 0
        rid = str(cell("id")).strip() if cell("id") else f"{concept_node_id}-{uuid.uuid4().hex[:6]}"
        try:
            items.append(ItemIn(
                item_id=rid, exam_code=exam, section_key=section_key, concept_node_id=concept_node_id,
                difficulty_d=diff, format=ItemFormatIn.mcq, num_options=len(opts),
                options=opts or None, correct_answer=str(cell("correct")).strip(),
                stem=str(stem).strip(),
                solution=(str(cell("sol")).strip() if cell("sol") else None),
                usage_scope=UsageScopeIn(scope if scope in ("both", "mock_only", "practice_only") else "both"),
            ))
        except Exception as e:  # noqa: BLE001
            errors.append({"row": seq, "item_id": rid, "error": str(e)})
    return items, errors


@router.post("/concepts/{node_id}/items/upload-xlsx")
async def upload_concept_quiz(node_id: str, file: UploadFile = File(...),
                              scope: str = "both", db: Session = Depends(get_db)) -> dict:
    """Bulk-upload a subtopic's quiz from a simple .xlsx (Question / options / Answer / Difficulty).
    Exam, section and concept come from the subtopic itself, so the sheet only carries questions.
    scope: 'both' (practice + mocks, default), 'practice_only', or 'mock_only'."""
    n = db.get(models.KnowledgeNode, node_id)
    if n is None or n.kind != "concept":
        raise HTTPException(404, f"no concept {node_id!r}")
    sec = db.get(models.Section, n.section_id)
    data = await file.read()
    items, parse_errors = _parse_quiz_xlsx(data, n.exam_code, sec.key if sec else None, node_id, scope)
    if not items:
        raise HTTPException(400, "no questions found (need a 'Question text' and 'Correct answer' column)")
    rep = ingestion.ingest_items(db, items)
    out = rep.model_dump() if hasattr(rep, "model_dump") else dict(rep)
    out["parse_errors"] = parse_errors
    return out


# ═══════════════════════ students: enrolment & management ═══════════════════════
class EnrollIn(BaseModel):
    exam_code: str


class StudentIn(BaseModel):
    name: str = ""
    email: str
    phone: Optional[str] = ""
    status: str = "active"
    regType: str = "registered"
    purchasedCourse: Optional[str] = None
    progress: int = 0
    verified: bool = False
    payment: Optional[dict] = None
    exams: list[str] = Field(default_factory=list)


class PaymentIn(BaseModel):
    status: Optional[str] = None
    amount: Optional[int] = None
    method: Optional[str] = None
    autoVerify: bool = False


class EnrollmentsIn(BaseModel):
    exams: list[str] = Field(default_factory=list)


def _account(db: Session, sid: str) -> "models.Account":
    try:
        a = db.get(models.Account, uuid.UUID(sid))
    except (ValueError, TypeError):
        a = None
    if a is None:
        raise HTTPException(404, "no such student")
    return a


def _ensure_profile(db: Session, account_id) -> "models.StudentProfile":
    p = db.get(models.StudentProfile, account_id)
    if p is None:
        p = models.StudentProfile(account_id=account_id)
        db.add(p)
        db.flush()
    return p


def _sync_enrollments(db: Session, account_id, exams) -> None:
    """Make the student's entitlements exactly match `exams` (add missing, remove extra)."""
    want = {e for e in (exams or []) if db.get(models.Exam, e) is not None}
    have = {e.exam_code: e for e in db.scalars(
        select(models.Entitlement).where(models.Entitlement.account_id == account_id)).all()}
    for ex in want - set(have):
        db.add(models.Entitlement(account_id=account_id, exam_code=ex, status="active"))
    for ex, ent in have.items():
        if ex not in want:
            db.delete(ent)


def _student_out(db: Session, a: "models.Account", admin_ids=None, *,
                 profile=None, ents=None, prefetched=False) -> dict:
    """The admin console's exact student shape (camelCase).

    Pass prefetched=True with `profile`/`ents` (from a batched query) to avoid a per-student round
    trip — critical over a remote DB, where the N+1 pattern is the main source of slowness."""
    if admin_ids is None:
        admin_ids = {x.account_id for x in db.scalars(select(models.AdminUser)).all()}
    p = profile if prefetched else db.get(models.StudentProfile, a.id)
    ents = ents if prefetched else db.scalars(
        select(models.Entitlement).where(models.Entitlement.account_id == a.id)).all()
    pay = (p.payment if (p and p.payment) else {}) or {}
    joined = (p.joined if (p and p.joined) else None) or (
        a.created_at.date().isoformat() if getattr(a, "created_at", None) else None)
    return {
        "id": str(a.id),
        "name": a.display_name or (a.email.split("@")[0] if a.email else ""),
        "email": a.email,
        "phone": (p.phone if p else "") or "",
        "exams": [e.exam_code for e in ents],
        "status": (p.status if p else "active") or "active",
        "regType": (p.reg_type if p else "registered") or "registered",
        "purchasedCourse": (p.purchased_course if p else None),
        "verified": bool(p.verified) if p else False,
        "lastLogin": (p.last_login if (p and p.last_login) else None) or "—",
        "progress": (p.progress if p else 0) or 0,
        "payment": {
            "status": pay.get("status"),
            "amount": pay.get("amount", 0),
            "method": pay.get("method"),
            "date": pay.get("date"),
        },
        "role": (p.role if p else "student") or "student",
        "joined": joined,
        "isAdmin": a.id in admin_ids,
    }


@router.get("/students")
def list_students(q: Optional[str] = None, limit: int = 200, db: Session = Depends(get_db)) -> dict:
    query = select(models.Account)
    if q:
        query = query.where(func.lower(models.Account.email).like(f"%{q.lower()}%"))
    accts = db.scalars(query.limit(limit)).all()
    ids = [a.id for a in accts]
    admin_ids = {x.account_id for x in db.scalars(select(models.AdminUser)).all()}
    # Batch the per-student data into 2 queries instead of 2*N (the N+1 that made this slow on Neon).
    profiles = {}
    ents_by: dict = {}
    if ids:
        for p in db.scalars(select(models.StudentProfile).where(models.StudentProfile.account_id.in_(ids))).all():
            profiles[p.account_id] = p
        for e in db.scalars(select(models.Entitlement).where(models.Entitlement.account_id.in_(ids))).all():
            ents_by.setdefault(e.account_id, []).append(e)
    out = [
        _student_out(db, a, admin_ids, profile=profiles.get(a.id),
                     ents=ents_by.get(a.id, []), prefetched=True)
        for a in accts
        if not (profiles.get(a.id) and profiles[a.id].deleted)
    ]
    return {"count": len(out), "students": out}


@router.get("/students/{sid}")
def student_detail(sid: str, db: Session = Depends(get_db)) -> dict:
    a = _account(db, sid)
    n_resp = db.scalar(select(func.count()).select_from(models.Response).where(
        models.Response.learner_id == a.id)) or 0
    out = _student_out(db, a)
    out["responses"] = int(n_resp)
    return out


@router.post("/students")
def create_student(body: StudentIn, db: Session = Depends(get_db)) -> dict:
    email = (body.email or "").strip().lower()
    if not email:
        raise HTTPException(400, "email is required")
    if db.scalar(select(models.Account).where(func.lower(models.Account.email) == email)):
        raise HTTPException(409, f"an account with email {email!r} already exists")
    a = models.Account(email=email, display_name=(body.name or "").strip() or email.split("@")[0])
    db.add(a)
    db.flush()
    db.add(models.StudentProfile(
        account_id=a.id, phone=(body.phone or ""), status=body.status, reg_type=body.regType,
        purchased_course=body.purchasedCourse, verified=body.verified, progress=body.progress,
        payment=(body.payment or {}), role="student", joined=datetime.now().date().isoformat(),
    ))
    _sync_enrollments(db, a.id, body.exams)
    db.commit()
    return _student_out(db, a)


@router.put("/students/{sid}")
def update_student(sid: str, body: StudentIn, db: Session = Depends(get_db)) -> dict:
    a = _account(db, sid)
    if body.name:
        a.display_name = body.name.strip()
    new_email = (body.email or "").strip().lower()
    if new_email and new_email != a.email:
        if db.scalar(select(models.Account).where(func.lower(models.Account.email) == new_email)):
            raise HTTPException(409, f"an account with email {new_email!r} already exists")
        a.email = new_email
    p = _ensure_profile(db, a.id)
    p.phone = body.phone or ""
    p.status = body.status
    p.reg_type = body.regType
    p.purchased_course = body.purchasedCourse
    p.progress = body.progress
    p.verified = body.verified
    if body.payment is not None:
        p.payment = body.payment
    _sync_enrollments(db, a.id, body.exams)
    db.commit()
    return _student_out(db, a)


@router.delete("/students/{sid}")
def delete_student(sid: str, db: Session = Depends(get_db)) -> dict:
    """Soft-delete: hides the student from the roster but preserves their learning history
    (responses, ability estimates) so nothing is silently destroyed and FKs stay intact."""
    a = _account(db, sid)
    p = _ensure_profile(db, a.id)
    p.deleted = True
    db.commit()
    return {"ok": True, "deleted": str(a.id)}


@router.post("/students/{sid}/verify")
def verify_student(sid: str, db: Session = Depends(get_db)) -> dict:
    a = _account(db, sid)
    p = _ensure_profile(db, a.id)
    p.verified = not p.verified
    db.commit()
    return _student_out(db, a)


@router.post("/students/{sid}/payment")
def set_student_payment(sid: str, body: PaymentIn, db: Session = Depends(get_db)) -> dict:
    a = _account(db, sid)
    p = _ensure_profile(db, a.id)
    pay = dict(p.payment or {})
    pay["status"] = body.status
    if body.amount is not None:
        pay["amount"] = body.amount
    if body.method:
        pay["method"] = body.method
    if not pay.get("date"):
        pay["date"] = datetime.now().date().isoformat()
    p.payment = pay
    if body.autoVerify and body.status == "successful":
        p.verified = True
    db.commit()
    return _student_out(db, a)


@router.put("/students/{sid}/enrollments")
def set_student_enrollments(sid: str, body: EnrollmentsIn, db: Session = Depends(get_db)) -> dict:
    a = _account(db, sid)
    _sync_enrollments(db, a.id, body.exams)
    db.commit()
    return _student_out(db, a)


@router.post("/students/{sid}/deregister")
def deregister_student(sid: str, body: EnrollIn, db: Session = Depends(get_db)) -> dict:
    """Remove a student from a course (revokes their entitlement for that exam)."""
    a = _account(db, sid)
    ents = db.scalars(select(models.Entitlement).where(
        models.Entitlement.account_id == a.id,
        models.Entitlement.exam_code == body.exam_code)).all()
    if not ents:
        raise HTTPException(404, f"student is not enrolled in {body.exam_code}")
    for e in ents:
        db.delete(e)
    db.commit()
    return {"ok": True, "deregistered": body.exam_code}


@router.post("/students/{sid}/enroll")
def enroll_student(sid: str, body: EnrollIn, db: Session = Depends(get_db)) -> dict:
    """Enrol a student into a course (grants/reactivates their entitlement)."""
    a = _account(db, sid)
    if db.get(models.Exam, body.exam_code) is None:
        raise HTTPException(404, f"no exam {body.exam_code!r}")
    ent = db.scalar(select(models.Entitlement).where(
        models.Entitlement.account_id == a.id, models.Entitlement.exam_code == body.exam_code))
    if ent:
        ent.status = "active"
    else:
        db.add(models.Entitlement(account_id=a.id, exam_code=body.exam_code, status="active"))
    db.commit()
    return {"ok": True, "enrolled": body.exam_code}


# ═══════════════════════════════ coupons / discount codes ═══════════════════════════════
class CouponIn(BaseModel):
    code: str
    type: str                       # 'percentage' | 'fixed'
    value: int = 0
    maxTotal: int = 0
    maxPerUser: int = 0
    minPurchase: int = 0
    maxDiscount: int = 0
    validFrom: Optional[str] = None
    validUntil: Optional[str] = None
    description: Optional[str] = ""
    attempt: str = "all"
    courses: list[str] = Field(default_factory=list)


def _coupon_out(c: "models.Coupon") -> dict:
    """Serialize to the admin console's exact camelCase shape."""
    return {
        "id": c.id, "code": c.code, "type": c.type, "value": c.value,
        "maxTotal": c.max_total, "maxPerUser": c.max_per_user,
        "minPurchase": c.min_purchase, "maxDiscount": c.max_discount,
        "validFrom": c.valid_from, "validUntil": c.valid_until,
        "description": c.description or "", "attempt": c.attempt,
        "courses": c.courses or [], "used": c.used, "status": c.status,
    }


@router.get("/coupons")
def list_coupons(db: Session = Depends(get_db)) -> dict:
    cs = db.scalars(select(models.Coupon).order_by(models.Coupon.created_at.desc())).all()
    return {"count": len(cs), "coupons": [_coupon_out(c) for c in cs]}


@router.post("/coupons")
def create_coupon(body: CouponIn, db: Session = Depends(get_db)) -> dict:
    code = "".join((body.code or "").split()).upper()   # strip ALL whitespace (no spaces in codes)
    if not code:
        raise HTTPException(400, "coupon code is required")
    if body.type not in ("percentage", "fixed"):
        raise HTTPException(400, "type must be 'percentage' or 'fixed'")
    if db.scalar(select(models.Coupon).where(func.upper(models.Coupon.code) == code)):
        raise HTTPException(409, f"a coupon with code {code!r} already exists")
    c = models.Coupon(
        id=uuid.uuid4().hex, code=code, type=body.type, value=body.value,
        max_total=body.maxTotal, max_per_user=body.maxPerUser, min_purchase=body.minPurchase,
        max_discount=body.maxDiscount, valid_from=(body.validFrom or None),
        valid_until=(body.validUntil or None), description=(body.description or ""),
        attempt=(body.attempt or "all"), courses=list(body.courses or []), used=0, status="active",
    )
    db.add(c)
    db.commit()
    return _coupon_out(c)


@router.put("/coupons/{cid}")
def update_coupon(cid: str, body: CouponIn, db: Session = Depends(get_db)) -> dict:
    c = db.get(models.Coupon, cid)
    if c is None:
        raise HTTPException(404, "no such coupon")
    code = "".join((body.code or "").split()).upper()   # strip ALL whitespace (no spaces in codes)
    if not code:
        raise HTTPException(400, "coupon code is required")
    if code != c.code and db.scalar(select(models.Coupon).where(func.upper(models.Coupon.code) == code)):
        raise HTTPException(409, f"a coupon with code {code!r} already exists")
    c.code, c.type, c.value = code, body.type, body.value
    c.max_total, c.max_per_user = body.maxTotal, body.maxPerUser
    c.min_purchase, c.max_discount = body.minPurchase, body.maxDiscount
    c.valid_from, c.valid_until = (body.validFrom or None), (body.validUntil or None)
    c.description, c.attempt = (body.description or ""), (body.attempt or "all")
    c.courses = list(body.courses or [])
    db.commit()
    return _coupon_out(c)


@router.post("/coupons/{cid}/toggle")
def toggle_coupon(cid: str, db: Session = Depends(get_db)) -> dict:
    c = db.get(models.Coupon, cid)
    if c is None:
        raise HTTPException(404, "no such coupon")
    c.status = "inactive" if c.status == "active" else "active"
    db.commit()
    return _coupon_out(c)


@router.delete("/coupons/{cid}")
def delete_coupon(cid: str, db: Session = Depends(get_db)) -> dict:
    c = db.get(models.Coupon, cid)
    if c is None:
        raise HTTPException(404, "no such coupon")
    code = c.code
    db.delete(c)
    db.commit()
    return {"ok": True, "deleted": code}


# ═══════════════════════════ study materials (PDF / slides / notes) ═══════════════════════════
_MAX_MATERIAL_BYTES = 25 * 1024 * 1024  # 25 MB cap for DB-stored files


def _size_label(n: int) -> str:
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        return f"{n / 1024:.0f} KB"
    return f"{n / (1024 * 1024):.1f} MB"


def _material_out(m: "models.Material") -> dict:
    return {
        "id": m.id, "title": m.title, "filename": m.filename, "contentType": m.content_type,
        "size": m.size_bytes, "sizeLabel": _size_label(m.size_bytes),
    }


@router.get("/concepts/{node_id}/materials")
def list_materials(node_id: str, db: Session = Depends(get_db)) -> dict:
    if db.get(models.KnowledgeNode, node_id) is None:
        raise HTTPException(404, "no such concept")
    ms = db.scalars(select(models.Material).where(models.Material.node_id == node_id)
                    .order_by(models.Material.created_at)).all()
    return {"count": len(ms), "materials": [_material_out(m) for m in ms]}


@router.post("/concepts/{node_id}/materials")
async def upload_material(node_id: str, file: UploadFile = File(...),
                          title: Optional[str] = Form(None), db: Session = Depends(get_db)) -> dict:
    if db.get(models.KnowledgeNode, node_id) is None:
        raise HTTPException(404, "no such concept")
    data = await file.read()
    if not data:
        raise HTTPException(400, "the uploaded file is empty")
    if len(data) > _MAX_MATERIAL_BYTES:
        raise HTTPException(413, f"file is too large (max {_size_label(_MAX_MATERIAL_BYTES)})")
    m = models.Material(
        id=uuid.uuid4().hex, node_id=node_id,
        title=(title or file.filename or "Untitled"),
        filename=(file.filename or "material"),
        content_type=(file.content_type or "application/octet-stream"),
        size_bytes=len(data), data=data,
    )
    db.add(m)
    db.commit()
    return _material_out(m)


@router.get("/materials/{mid}/download")
def admin_download_material(mid: str, db: Session = Depends(get_db)):
    """Stream a material's bytes for admin preview. Admin-gated at the router level and deliberately
    NOT entitlement-gated — an admin can view any uploaded material regardless of course access."""
    m = db.get(models.Material, mid)
    if m is None:
        raise HTTPException(404, "no such material")
    return Response(content=m.data, media_type=m.content_type or "application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{m.filename}"'})


@router.delete("/materials/{mid}")
def delete_material(mid: str, db: Session = Depends(get_db)) -> dict:
    m = db.get(models.Material, mid)
    if m is None:
        raise HTTPException(404, "no such material")
    db.delete(m)
    db.commit()
    return {"ok": True, "deleted": mid}


# ═══════════════════════════ question images (media) ═══════════════════════════
_MAX_IMAGE_BYTES = 8 * 1024 * 1024  # 8 MB per image


@router.post("/media")
async def upload_media(files: list[UploadFile] = File(...), db: Session = Depends(get_db)) -> dict:
    """Bulk-upload question images. Each file is keyed by its filename minus extension (= the
    question/item id), e.g. `q123.png` -> key `q123`. Re-uploading a key replaces it. The image then
    auto-attaches to the matching question wherever it appears (mocks, diagnostic, practice)."""
    from ..services.media import key_from_filename

    out = []
    for f in files:
        data = await f.read()
        if not data:
            continue
        if len(data) > _MAX_IMAGE_BYTES:
            raise HTTPException(413, f"{f.filename}: too large (max {_size_label(_MAX_IMAGE_BYTES)})")
        key = key_from_filename(f.filename)
        if not key:
            continue
        existing = db.get(models.MediaAsset, key)
        if existing:
            existing.data = data
            existing.content_type = f.content_type or existing.content_type
            existing.size_bytes = len(data)
        else:
            db.add(models.MediaAsset(key=key, content_type=(f.content_type or "image/png"),
                                     size_bytes=len(data), data=data))
        out.append({"key": key, "url": f"/media/{key}", "size": len(data)})
    db.commit()
    return {"uploaded": out, "count": len(out)}


@router.get("/media")
def list_media(db: Session = Depends(get_db)) -> dict:
    rows = db.query(
        models.MediaAsset.key, models.MediaAsset.content_type,
        models.MediaAsset.size_bytes, models.MediaAsset.created_at,
    ).order_by(models.MediaAsset.created_at.desc()).all()
    return {
        "items": [{"key": k, "url": f"/media/{k}", "contentType": ct, "size": sz,
                   "createdAt": ca.isoformat() if ca else None} for (k, ct, sz, ca) in rows],
        "count": len(rows),
    }


@router.delete("/media/{key}")
def delete_media(key: str, db: Session = Depends(get_db)) -> dict:
    m = db.get(models.MediaAsset, key)
    if m is None:
        raise HTTPException(404, "no such image")
    db.delete(m)
    db.commit()
    return {"ok": True, "deleted": key}


# ═══════════════════════════ contact-us submissions ═══════════════════════════
def _contact_out(m: "models.ContactMessage") -> dict:
    return {"id": m.id, "firstName": m.first_name, "lastName": m.last_name,
            "phone": m.phone, "email": m.email, "message": m.message,
            "handled": bool(m.handled),
            "createdAt": m.created_at.isoformat() if m.created_at else None}


@router.get("/contact")
def list_contact(db: Session = Depends(get_db)) -> dict:
    rows = db.query(models.ContactMessage).order_by(models.ContactMessage.created_at.desc()).all()
    return {"items": [_contact_out(m) for m in rows], "count": len(rows),
            "unhandled": sum(1 for m in rows if not m.handled)}


@router.patch("/contact/{cid}")
def set_contact_handled(cid: str, handled: bool = True, db: Session = Depends(get_db)) -> dict:
    m = db.get(models.ContactMessage, cid)
    if m is None:
        raise HTTPException(404, "no such message")
    m.handled = handled
    db.commit()
    return _contact_out(m)


@router.delete("/contact/{cid}")
def delete_contact(cid: str, db: Session = Depends(get_db)) -> dict:
    m = db.get(models.ContactMessage, cid)
    if m is None:
        raise HTTPException(404, "no such message")
    db.delete(m)
    db.commit()
    return {"ok": True, "deleted": cid}


# ═══════════════════════════ mock builder (fixed-form mocks) ═══════════════════════════
class MockCreateIn(BaseModel):
    type: str                       # sectional | full
    name: str
    exam: str
    negative: int = 0
    duration: int = 0
    scoringMarks: int = 1
    scoringNeg: int = 0
    instructions: Optional[str] = ""


class MockConfigIn(BaseModel):
    name: Optional[str] = None
    negative: Optional[int] = None
    duration: Optional[int] = None
    scoringMarks: Optional[int] = None
    scoringNeg: Optional[int] = None
    instructions: Optional[str] = None


class MockStructureIn(BaseModel):
    sections: list = Field(default_factory=list)  # full sections tree with embedded questions


def _mock_out(m: "models.Mock") -> dict:
    secs = m.sections or []
    tq = sum(len(s.get("questions", [])) for s in secs)
    tm = sum(int(s.get("time", 0) or 0) for s in secs)
    return {
        "id": m.id, "exam": m.exam_code, "type": m.type, "name": m.name, "status": m.status,
        "negative": m.negative, "duration": m.duration,
        "scoringMarks": m.scoring_marks, "scoringNeg": m.scoring_neg,
        "instructions": m.instructions or "", "sections": secs,
        "totalQuestions": tq, "totalTime": tm,
        "attempts": [],  # results need a student mock-taking flow (a later phase)
    }


def _mock_or_404(db: Session, mid: str) -> "models.Mock":
    m = db.get(models.Mock, mid)
    if m is None:
        raise HTTPException(404, "no such mock")
    return m


@router.get("/mocks")
def list_mocks(exam: str, type: Optional[str] = None, db: Session = Depends(get_db)) -> dict:
    q = select(models.Mock).where(models.Mock.exam_code == exam)
    if type:
        q = q.where(models.Mock.type == type)
    ms = db.scalars(q.order_by(models.Mock.created_at)).all()
    return {"count": len(ms), "mocks": [_mock_out(m) for m in ms]}


@router.post("/mocks")
def create_mock(body: MockCreateIn, db: Session = Depends(get_db)) -> dict:
    if body.type not in ("sectional", "full", "diagnostic"):
        raise HTTPException(400, "type must be 'sectional', 'full', or 'diagnostic'")
    if db.get(models.Exam, body.exam) is None:
        raise HTTPException(404, f"no exam {body.exam!r}")
    keys = [s.key for s in db.scalars(select(models.Section).where(
        models.Section.exam_code == body.exam).order_by(models.Section.key)).all()]
    if body.type == "sectional":
        first = keys[0] if keys else "Section 1"
        sections = [{"id": uuid.uuid4().hex, "name": first, "time": 40, "numQuestions": 0, "questions": []}]
    else:
        sections = ([{"id": uuid.uuid4().hex, "name": k, "questions": []} for k in keys]
                    or [{"id": uuid.uuid4().hex, "name": "Section 1", "questions": []}])
    m = models.Mock(
        id=uuid.uuid4().hex, exam_code=body.exam, type=body.type, name=body.name, status="draft",
        negative=body.negative, duration=body.duration, scoring_marks=body.scoringMarks,
        scoring_neg=body.scoringNeg, instructions=(body.instructions or ""), sections=sections,
    )
    db.add(m)
    db.commit()
    return _mock_out(m)


@router.get("/mocks/{mid}")
def get_mock(mid: str, db: Session = Depends(get_db)) -> dict:
    return _mock_out(_mock_or_404(db, mid))


@router.put("/mocks/{mid}")
def update_mock_config(mid: str, body: MockConfigIn, db: Session = Depends(get_db)) -> dict:
    m = _mock_or_404(db, mid)
    if body.name is not None:
        m.name = body.name
    if body.negative is not None:
        m.negative = body.negative
    if body.duration is not None:
        m.duration = body.duration
    if body.scoringMarks is not None:
        m.scoring_marks = body.scoringMarks
    if body.scoringNeg is not None:
        m.scoring_neg = body.scoringNeg
    if body.instructions is not None:
        m.instructions = body.instructions
    db.commit()
    return _mock_out(m)


@router.put("/mocks/{mid}/structure")
def set_mock_structure(mid: str, body: MockStructureIn, db: Session = Depends(get_db)) -> dict:
    """Replace the whole sections/questions tree. The console mutates the mock in memory
    (add/edit/move/delete sections and questions, bulk import) then persists the result here."""
    m = _mock_or_404(db, mid)
    secs = body.sections or []
    for s in secs:
        if not isinstance(s, dict):
            raise HTTPException(400, "each section must be an object")
        s.setdefault("id", uuid.uuid4().hex)
        for qz in s.get("questions", []) or []:
            if isinstance(qz, dict):
                qz.setdefault("id", uuid.uuid4().hex)
    m.sections = secs
    db.commit()
    return _mock_out(m)


@router.post("/mocks/{mid}/publish")
def toggle_mock_publish(mid: str, db: Session = Depends(get_db)) -> dict:
    m = _mock_or_404(db, mid)
    m.status = "draft" if m.status == "published" else "published"
    db.commit()
    return _mock_out(m)


@router.delete("/mocks/{mid}")
def delete_mock(mid: str, db: Session = Depends(get_db)) -> dict:
    m = _mock_or_404(db, mid)
    # Remove every student's attempt data for this mock, so attempt counts / best-last cards don't
    # outlive the mock (a deleted mock leaves no orphaned attempts for any student).
    a = db.execute(delete(models.MockAttempt).where(models.MockAttempt.mock_id == mid)).rowcount
    db.execute(delete(models.DiagnosticAttempt).where(models.DiagnosticAttempt.mock_id == mid))
    db.delete(m)
    db.commit()
    return {"ok": True, "deleted": mid, "attempts_removed": a}
